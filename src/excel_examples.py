"""
Excel operations examples.
Demonstrates Excel file reading, writing, and data manipulation.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from config import DATA_DIR, SAMPLE_EXCEL_FILE, DEFAULT_SHEET_NAME


class ExcelProcessor:
    """Simple Excel file processor."""
    
    def __init__(self):
        """Initialize Excel processor."""
        # Ensure data directory exists
        os.makedirs(DATA_DIR, exist_ok=True)
        print(f"📁 Excel processor initialized. Data directory: {DATA_DIR}")
    
    def create_sample_data(self):
        """Create sample Excel file with test data."""
        sample_data = {
            'ID': [1, 2, 3, 4, 5],
            'Name': ['John Doe', 'Jane Smith', 'Bob Johnson', 'Alice Brown', 'Charlie Wilson'],
            'Age': [25, 30, 35, 28, 42],
            'Department': ['IT', 'HR', 'Finance', 'IT', 'Marketing'],
            'Salary': [50000, 55000, 60000, 52000, 65000],
            'Start Date': ['2023-01-15', '2022-03-20', '2021-07-10', '2023-02-01', '2020-11-05']
        }
        
        try:
            df = pd.DataFrame(sample_data)
            df['Start Date'] = pd.to_datetime(df['Start Date'])
            df.to_excel(SAMPLE_EXCEL_FILE, index=False, sheet_name=DEFAULT_SHEET_NAME)
            print(f"✅ Created sample Excel file: {SAMPLE_EXCEL_FILE}")
            return True
        except Exception as e:
            print(f"❌ Error creating sample Excel file: {e}")
            return False
    
    def read_excel_pandas(self, file_path=None):
        """Read Excel file using pandas."""
        if file_path is None:
            file_path = SAMPLE_EXCEL_FILE
        
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"❌ File not found: {file_path}")
                return None
            
            # Read Excel file
            df = pd.read_excel(file_path, sheet_name=DEFAULT_SHEET_NAME)
            print(f"✅ Read Excel file: {file_path}")
            print(f"📊 Shape: {df.shape} (rows: {df.shape[0]}, columns: {df.shape[1]})")
            print(f"📋 Columns: {list(df.columns)}")
            print("\n🔍 First 5 rows:")
            print(df.head())
            
            return df
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return None
    
    def write_excel_pandas(self, data, file_path, sheet_name=DEFAULT_SHEET_NAME):
        """Write data to Excel using pandas."""
        try:
            if isinstance(data, dict):
                df = pd.DataFrame(data)
            elif isinstance(data, pd.DataFrame):
                df = data
            else:
                print("❌ Data must be a dictionary or DataFrame")
                return False
            
            df.to_excel(file_path, index=False, sheet_name=sheet_name)
            print(f"✅ Written Excel file: {file_path}")
            print(f"📊 Shape: {df.shape}")
            return True
        except Exception as e:
            print(f"❌ Error writing Excel file: {e}")
            return False
    
    def read_excel_openpyxl(self, file_path=None):
        """Read Excel file using openpyxl (more control)."""
        if file_path is None:
            file_path = SAMPLE_EXCEL_FILE
        
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"❌ File not found: {file_path}")
                return None
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            print(f"✅ Loaded Excel file: {file_path}")
            print(f"📋 Sheet name: {ws.title}")
            print(f"📊 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Read data
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            print("\n🔍 First 3 rows:")
            for i, row in enumerate(data[:3]):
                print(f"Row {i+1}: {row}")
            
            wb.close()
            return data
        except Exception as e:
            print(f"❌ Error reading Excel file with openpyxl: {e}")
            return None
    
    def write_excel_formatted(self, data, file_path):
        """Write Excel file with formatting using openpyxl."""
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Formatted Data"
            
            # Convert data to DataFrame if needed
            if isinstance(data, dict):
                df = pd.DataFrame(data)
            elif isinstance(data, pd.DataFrame):
                df = data
            else:
                print("❌ Data must be a dictionary or DataFrame")
                return False
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            for cell in ws[1]:  # First row (headers)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            print(f"✅ Created formatted Excel file: {file_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error creating formatted Excel file: {e}")
            return False
    
    def analyze_data(self, df):
        """Perform basic data analysis on DataFrame."""
        if df is None:
            print("❌ No data to analyze")
            return None
        
        try:
            print("\n📈 Data Analysis:")
            print("=" * 30)
            
            # Basic info
            print(f"📊 Dataset shape: {df.shape}")
            print(f"📋 Columns: {list(df.columns)}")
            
            # Data types
            print("\n🔢 Data types:")
            print(df.dtypes)
            
            # Numeric columns analysis
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                print(f"\n📊 Numeric columns summary:")
                print(df[numeric_cols].describe())
            
            # Missing values
            missing = df.isnull().sum()
            if missing.sum() > 0:
                print(f"\n❓ Missing values:")
                print(missing[missing > 0])
            else:
                print("\n✅ No missing values found")
            
            # Categorical analysis
            categorical_cols = df.select_dtypes(include=['object']).columns
            if len(categorical_cols) > 0:
                print(f"\n📝 Categorical columns:")
                for col in categorical_cols:
                    unique_count = df[col].nunique()
                    print(f"  {col}: {unique_count} unique values")
                    if unique_count <= 10:  # Show values if not too many
                        print(f"    Values: {list(df[col].unique())}")
            
            return {
                'shape': df.shape,
                'columns': list(df.columns),
                'numeric_summary': df.select_dtypes(include=['number']).describe().to_dict() if len(numeric_cols) > 0 else {},
                'missing_values': missing.to_dict(),
                'categorical_summary': {col: df[col].nunique() for col in categorical_cols}
            }
            
        except Exception as e:
            print(f"❌ Error analyzing data: {e}")
            return None
    
    def filter_and_transform(self, df):
        """Demonstrate data filtering and transformation."""
        if df is None:
            print("❌ No data to filter")
            return None
        
        try:
            print("\n🔄 Data Filtering and Transformation:")
            print("=" * 40)
            
            # Filter examples
            if 'Age' in df.columns:
                young_employees = df[df['Age'] < 30]
                print(f"👥 Employees under 30: {len(young_employees)}")
                print(young_employees[['Name', 'Age', 'Department']])
            
            if 'Department' in df.columns:
                it_dept = df[df['Department'] == 'IT']
                print(f"\n💻 IT Department employees: {len(it_dept)}")
                print(it_dept[['Name', 'Age', 'Salary']])
            
            # Group by analysis
            if 'Department' in df.columns and 'Salary' in df.columns:
                dept_salary = df.groupby('Department')['Salary'].agg(['mean', 'count']).round(2)
                print(f"\n💰 Average salary by department:")
                print(dept_salary)
            
            # Add calculated columns
            if 'Salary' in df.columns:
                df_copy = df.copy()
                df_copy['Annual Bonus'] = df_copy['Salary'] * 0.1
                df_copy['Monthly Salary'] = df_copy['Salary'] / 12
                
                print(f"\n💡 Added calculated columns:")
                print(df_copy[['Name', 'Salary', 'Annual Bonus', 'Monthly Salary']].head())
                
                return df_copy
            
            return df
            
        except Exception as e:
            print(f"❌ Error filtering/transforming data: {e}")
            return None


def demo_excel_operations():
    """Demonstrate Excel operations."""
    print("\n📊 Excel Operations Demo")
    print("=" * 30)
    
    processor = ExcelProcessor()
    
    # Create sample data
    processor.create_sample_data()
    
    # Read with pandas
    print("\n📖 Reading with pandas:")
    df = processor.read_excel_pandas()
    
    # Read with openpyxl
    print("\n📖 Reading with openpyxl:")
    processor.read_excel_openpyxl()
    
    # Analyze data
    if df is not None:
        analysis = processor.analyze_data(df)
        
        # Filter and transform
        transformed_df = processor.filter_and_transform(df)
        
        # Write new files
        output_file = os.path.join(DATA_DIR, 'processed_data.xlsx')
        processor.write_excel_pandas(transformed_df, output_file)
        
        # Write formatted file
        formatted_file = os.path.join(DATA_DIR, 'formatted_data.xlsx')
        processor.write_excel_formatted(transformed_df, formatted_file)


def demo_multiple_sheets():
    """Demonstrate working with multiple Excel sheets."""
    print("\n📚 Multiple Sheets Demo")
    print("=" * 30)
    
    try:
        # Create data for different sheets
        employees_data = {
            'ID': [1, 2, 3, 4],
            'Name': ['John', 'Jane', 'Bob', 'Alice'],
            'Department': ['IT', 'HR', 'Finance', 'IT']
        }
        
        departments_data = {
            'Department': ['IT', 'HR', 'Finance', 'Marketing'],
            'Budget': [100000, 80000, 120000, 90000],
            'Head': ['Smith', 'Johnson', 'Williams', 'Brown']
        }
        
        # Create workbook with multiple sheets
        multi_sheet_file = os.path.join(DATA_DIR, 'multi_sheet_data.xlsx')
        
        with pd.ExcelWriter(multi_sheet_file, engine='openpyxl') as writer:
            pd.DataFrame(employees_data).to_excel(writer, sheet_name='Employees', index=False)
            pd.DataFrame(departments_data).to_excel(writer, sheet_name='Departments', index=False)
        
        print(f"✅ Created multi-sheet Excel file: {multi_sheet_file}")
        
        # Read specific sheets
        employees_df = pd.read_excel(multi_sheet_file, sheet_name='Employees')
        departments_df = pd.read_excel(multi_sheet_file, sheet_name='Departments')
        
        print(f"\n📋 Employees sheet:")
        print(employees_df)
        
        print(f"\n🏢 Departments sheet:")
        print(departments_df)
        
        # Read all sheets
        all_sheets = pd.read_excel(multi_sheet_file, sheet_name=None)
        print(f"\n📚 All sheets: {list(all_sheets.keys())}")
        
    except Exception as e:
        print(f"❌ Error with multiple sheets: {e}")


if __name__ == "__main__":
    demo_excel_operations()
    demo_multiple_sheets()
